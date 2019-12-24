#!/usr/bin/python3
print()

password = "9826717511"
import cgi
import cgitb
cgitb.enable()
form = cgi.FieldStorage()
passcode = form.getvalue("pass")

if passcode == password:
	try:
		f = open("data/stockrecovery.txt",'r')
		lines = f.read().splitlines()
		dt = lines[-1]
		week = ""
		val = int(dt[8:10])
		if val <= 7:
			week = "week1"
		elif val >= 8 and val <= 14:
			week = "week2"
		elif val >= 15 and val <= 21:
			week = "week3"
		elif val >= 22:
			week = "week4"

		from openpyxl import Workbook, load_workbook
		wb = load_workbook("data/"+dt[:4]+"/"+dt[5:7]+"/"+week+"/data.xlsx")
		ws = wb["addstock"]
		p = ws.cell(row=ws.max_row,column=3).value
		d = ws.cell(row=ws.max_row,column=5).value
		if ws.max_row!=1:
			ws.delete_rows(ws.max_row)
		wb.save("data/"+dt[:4]+"/"+dt[5:7]+"/"+week+"/data.xlsx")
		
		wb = load_workbook("data/"+dt[:4]+"/"+dt[5:7]+"/data.xlsx")
		ws = wb["addstock"]
		if ws.max_row!=1:
			ws.delete_rows(ws.max_row)
		wb.save("data/"+dt[:4]+"/"+dt[5:7]+"/data.xlsx")
		
		wb = load_workbook("data/"+dt[:4]+"/data.xlsx")
		ws = wb["addstock"]
		if ws.max_row!=1:
			ws.delete_rows(ws.max_row)
		wb.save("data/"+dt[:4]+"/data.xlsx")
		
		import os
		petrol = open("data/petrol.txt","r")
		petrol = petrol.read()
		petrol = petrol[:-1]
		petrol = float(petrol) - float(p)
		petrol = str(petrol)
		os.system("echo "+petrol+" > data/petrol.txt")
		
		diesel = open("data/diesel.txt","r")
		diesel = diesel.read()
		diesel = diesel[:-1]
		diesel = float(diesel) - float(d)
		diesel = str(diesel)
		os.system("echo "+diesel+" > data/diesel.txt")
		
		os.system("rm data/stockrecovery.txt")
		for line in lines[:-1]:
			os.system("echo "+line+" >> data/stockrecovery.txt")
		
		print("Last stock record deleted successfully")
	except:
		print("Only 10 deletes are allowed")
else:
	print("Wrong PassCode")
