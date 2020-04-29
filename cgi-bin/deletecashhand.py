#!/usr/bin/python3
print()

password = "2870"
import cgi
import cgitb
cgitb.enable()
form = cgi.FieldStorage()
passcode = form.getvalue("pass")

if passcode == password:
	try:
		f = open("data/cashhandrecovery.txt",'r')
		lines = f.read().splitlines()
		dt = lines[-1]
		week = ""
		val = int(dt[8:10])
		if val <= 7:
			week = "week1"
		elif val >= 8 and val <= 14:
			week = "week2"
		elif val >= 15 and val <= 21:
			week = "week3"
		elif val >= 22:
			week = "week4"

		from openpyxl import Workbook, load_workbook
		wb = load_workbook("data/"+dt[:4]+"/"+dt[5:7]+"/"+week+"/data.xlsx")
		ws = wb["addcashhand"]
		ch = ws.cell(row=ws.max_row,column=11).value
		d = ws.cell(row=ws.max_row,column=12).value
		l = ws.cell(row=ws.max_row,column=13).value
		if ws.max_row!=1:
			ws.delete_rows(ws.max_row)
		wb.save("data/"+dt[:4]+"/"+dt[5:7]+"/"+week+"/data.xlsx")
		
		wb = load_workbook("data/"+dt[:4]+"/"+dt[5:7]+"/data.xlsx")
		ws = wb["addcashhand"]
		if ws.max_row!=1:
			ws.delete_rows(ws.max_row)
		wb.save("data/"+dt[:4]+"/"+dt[5:7]+"/data.xlsx")
		
		wb = load_workbook("data/"+dt[:4]+"/data.xlsx")
		ws = wb["addcashhand"]
		if ws.max_row!=1:
			ws.delete_rows(ws.max_row)
		wb.save("data/"+dt[:4]+"/data.xlsx")
		
		import os
		locker = open("data/locker.txt","r")
		locker = locker.read()
		locker = locker[:-1]
		locker = float(l) - float(ch) + float(d)
		locker = str(locker)
		os.system("echo "+locker+" > data/locker.txt")
		
		os.system("rm data/cashhandrecovery.txt")
		for line in lines[:-1]:
			os.system("echo "+line+" >> data/cashhandrecovery.txt")
		
		print("Last cashhand record deleted successfully")
	except:
		print("Only 10 deletes are allowed")
else:
	print("Wrong PassCode")
