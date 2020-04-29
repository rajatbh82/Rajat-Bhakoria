#!/usr/bin/python3
print()

password = "2870"
import cgi
import cgitb
cgitb.enable()
form = cgi.FieldStorage()
#week month or year
wmy = form.getvalue("wmy")
dt = form.getvalue("date")
passcode = form.getvalue("pass")

if passcode == password:
	if wmy == "week":
		week = ""
		val = int(dt[8:10])
		if val <= 7:
			week = "week1"
		elif val >= 8 and val <= 14:
			week = "week2"
		elif val >= 15 and val <= 21:
			week = "week3"
		elif val >= 22:
			week = "week4"
		
		try:
			from openpyxl import Workbook, load_workbook
			wb = load_workbook("data/"+dt[:4]+"/"+dt[5:7]+"/"+week+"/data.xlsx")
			addstock = wb["addstock"]
			addsale = wb["addsale"]
			
			try:
				wb.remove(wb["report"])
			except:
				print()
			wb.create_sheet("report",3)
			wb['report'].append(["Date","Invoice Number","Petrol Bought","Petrol Cost Price","Diesel Bought","Diesel Cost Price","Total Cost Price","Petrol in Stock","Petrol Sold","Petrol for Testing","Petrol Rate","Petrol Selling Price","Diesel in Stock","Diesel Sold","Diesel for Testing","Diesel Rate","Diesel Selling Price","Total Selling Price"])
			
			petrol = open("data/"+dt[:4]+"/"+dt[5:7]+"/"+week+"/openpetrol.txt")
			petrol = petrol.read()
			petrol = petrol[:-1]
			
			diesel = open("data/"+dt[:4]+"/"+dt[5:7]+"/"+week+"/opendiesel.txt")
			diesel = diesel.read()
			diesel = diesel[:-1]
			
			count = 2
			for i in range(2,addsale.max_row+1):
				lst = []
				lst.append(addsale.cell(row=i,column=1).value)
				if addsale.cell(row=i,column=1).value == addstock.cell(row=count,column=1).value :
					for j in range(2,addstock.max_column+1):
						lst.append(addstock.cell(row=count,column=j).value)
					lst.append(petrol)
					for j in range(2,6):
						lst.append(addsale.cell(row=i,column=j).value)
					lst.append(diesel)
					for j in range(6,addsale.max_column+1):
						lst.append(addsale.cell(row=i,column=j).value)
					petrol = str(float(petrol) + float(addstock.cell(row=count,column=3).value) - float(addsale.cell(row=i,column=2).value) - float(addsale.cell(row=i,column=3).value))
					diesel = str(float(diesel) + float(addstock.cell(row=count,column=5).value) - float(addsale.cell(row=i,column=6).value) - float(addsale.cell(row=i,column=7).value))
					count = count + 1
				else:
					lst.extend(["","","","","",""])
					lst.append(petrol)
					for j in range(2,6):
						lst.append(addsale.cell(row=i,column=j).value)
					lst.append(diesel)
					for j in range(6,addsale.max_column+1):
						lst.append(addsale.cell(row=i,column=j).value)
					petrol = str(float(petrol) - float(addsale.cell(row=i,column=2).value) - float(addsale.cell(row=i,column=3).value))
					diesel = str(float(diesel) - float(addsale.cell(row=i,column=6).value) - float(addsale.cell(row=i,column=7).value))
				wb["report"].append(lst)
			wb.save("data/"+dt[:4]+"/"+dt[5:7]+"/"+week+"/data.xlsx")
			print("<a href=\"data/"+dt[:4]+"/"+dt[5:7]+"/"+week+"/data.xlsx\" download>Download</a><br><table style='font-size:10px'><tr>")
			ws = wb["report"]
			for i in range(1,ws.max_column+1):
				print("<th>"+ws.cell(row=1,column=i).value+"</th>")
			print("</tr>")
			for i in range(2,ws.max_row+1):
				print("<tr>")
				for j in range(1,ws.max_column+1):
					print("<td>"+ws.cell(row=i,column=j).value+"</td>")
				print("</tr>")
			print("</table>")
		except:
			print("No record found for selected year")
	if wmy == "month":
		try:
			from openpyxl import Workbook, load_workbook
			wb = load_workbook("data/"+dt[:4]+"/"+dt[5:7]+"/data.xlsx")
			addstock = wb["addstock"]
			addsale = wb["addsale"]
			try:
				wb.remove(wb["report"])
			except:
				print()
			wb.create_sheet("report",3)
			wb['report'].append(["Date","Invoice Number","Petrol Bought","Petrol Cost Price","Diesel Bought","Diesel Cost Price","Total Cost Price","Petrol in Stock","Petrol Sold","Petrol for Testing","Petrol Rate","Petrol Selling Price","Diesel in Stock","Diesel Sold","Diesel for Testing","Diesel Rate","Diesel Selling Price","Total Selling Price"])
			
			petrol = open("data/"+dt[:4]+"/"+dt[5:7]+"/openpetrol.txt")
			petrol = petrol.read()
			petrol = petrol[:-1]
			
			diesel = open("data/"+dt[:4]+"/"+dt[5:7]+"/opendiesel.txt")
			diesel = diesel.read()
			diesel = diesel[:-1]
			
			count = 2
			for i in range(2,addsale.max_row+1):
				lst = []
				lst.append(addsale.cell(row=i,column=1).value)
				if addsale.cell(row=i,column=1).value == addstock.cell(row=count,column=1).value :
					for j in range(2,addstock.max_column+1):
						lst.append(addstock.cell(row=count,column=j).value)
					lst.append(petrol)
					for j in range(2,6):
						lst.append(addsale.cell(row=i,column=j).value)
					lst.append(diesel)
					for j in range(6,addsale.max_column+1):
						lst.append(addsale.cell(row=i,column=j).value)
					petrol = str(float(petrol) + float(addstock.cell(row=count,column=3).value) - float(addsale.cell(row=i,column=2).value))
					diesel = str(float(diesel) + float(addstock.cell(row=count,column=5).value) - float(addsale.cell(row=i,column=6).value))
					count = count + 1
				else:
					lst.extend(["","","","","",""])
					lst.append(petrol)
					for j in range(2,6):
						lst.append(addsale.cell(row=i,column=j).value)
					lst.append(diesel)
					for j in range(6,addsale.max_column+1):
						lst.append(addsale.cell(row=i,column=j).value)
					petrol = str(float(petrol) - float(addsale.cell(row=i,column=2).value))
					diesel = str(float(diesel) - float(addsale.cell(row=i,column=6).value))
				wb["report"].append(lst)
			wb.save("data/"+dt[:4]+"/"+dt[5:7]+"/data.xlsx")
			print("<a href=\"data/"+dt[:4]+"/"+dt[5:7]+"/data.xlsx\" download>Download</a><br><table style='font-size:10px'><tr>")
			ws = wb["report"]
			for i in range(1,ws.max_column+1):
				print("<th>"+ws.cell(row=1,column=i).value+"</th>")
			print("</tr>")
			for i in range(2,ws.max_row+1):
				print("<tr>")
				for j in range(1,ws.max_column+1):
					print("<td>"+ws.cell(row=i,column=j).value+"</td>")
				print("</tr>")
			print("</table>")
		except:
			print("No record found for selected month")
	if wmy == "year":
		try:
			from openpyxl import Workbook, load_workbook
			wb = load_workbook("data/"+dt[:4]+"/data.xlsx")
			addstock = wb["addstock"]
			addsale = wb["addsale"]
			try:
				wb.remove(wb["report"])
			except:
				print()
			wb.create_sheet("report",3)
			wb['report'].append(["Date","Invoice Number","Petrol Bought","Petrol Cost Price","Diesel Bought","Diesel Cost Price","Total Cost Price","Petrol in Stock","Petrol Sold","Petrol for Testing","Petrol Rate","Petrol Selling Price","Diesel in Stock","Diesel Sold","Diesel for Testing","Diesel Rate","Diesel Selling Price","Total Selling Price"])
			
			petrol = open("data/"+dt[:4]+"/openpetrol.txt")
			petrol = petrol.read()
			petrol = petrol[:-1]
			
			diesel = open("data/"+dt[:4]+"/opendiesel.txt")
			diesel = diesel.read()
			diesel = diesel[:-1]
			
			count = 2
			for i in range(2,addsale.max_row+1):
				lst = []
				lst.append(addsale.cell(row=i,column=1).value)
				if addsale.cell(row=i,column=1).value == addstock.cell(row=count,column=1).value :
					for j in range(2,addstock.max_column+1):
						lst.append(addstock.cell(row=count,column=j).value)
					lst.append(petrol)
					for j in range(2,6):
						lst.append(addsale.cell(row=i,column=j).value)
					lst.append(diesel)
					for j in range(6,addsale.max_column+1):
						lst.append(addsale.cell(row=i,column=j).value)
					petrol = str(float(petrol) + float(addstock.cell(row=count,column=3).value) - float(addsale.cell(row=i,column=2).value))
					diesel = str(float(diesel) + float(addstock.cell(row=count,column=5).value) - float(addsale.cell(row=i,column=6).value))
					count = count + 1
				else:
					lst.extend(["","","","","",""])
					lst.append(petrol)
					for j in range(2,6):
						lst.append(addsale.cell(row=i,column=j).value)
					lst.append(diesel)
					for j in range(6,addsale.max_column+1):
						lst.append(addsale.cell(row=i,column=j).value)
					petrol = str(float(petrol) - float(addsale.cell(row=i,column=2).value))
					diesel = str(float(diesel) - float(addsale.cell(row=i,column=6).value))
				wb["report"].append(lst)
			wb.save("data/"+dt[:4]+"/data.xlsx")
			print("<a href=\"data/"+dt[:4]+"/data.xlsx\" download>Download</a><br><table style='font-size:10px'><tr>")
			ws = wb["report"]
			for i in range(1,ws.max_column+1):
				print("<th>"+ws.cell(row=1,column=i).value+"</th>")
			print("</tr>")
			for i in range(2,ws.max_row+1):
				print("<tr>")
				for j in range(1,ws.max_column+1):
					print("<td>"+ws.cell(row=i,column=j).value+"</td>")
				print("</tr>")
			print("</table>")
		except:
			print("No record found for selected Year")
else:
	print("Wrong passcode")
