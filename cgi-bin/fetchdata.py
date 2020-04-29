#!/usr/bin/python3
print()


import cgi
import cgitb
cgitb.enable()
form = cgi.FieldStorage()
lastdate = "<span style=\"color:navy\">Date of last Record filled</span><table style='font-size:80%'><tr><th>Stock</th><th>Sale</th><th>Cashhand</th></tr>"
stock = ""
sale = ""
cash = ""
try:
	f = open("data/stockrecovery.txt",'r')
	lines = f.read().splitlines()
	dt = lines[-1]
	week = ""
	val = int(dt[8:10])
	if val <= 7:
		week = "week1"
	elif val >= 8 and val <= 14:
		week = "week2"
	elif val >= 15 and val <= 21:
		week = "week3"
	elif val >= 22:
		week = "week4"
	lastdate = lastdate + "<tr><td>"+dt+"</td>"
	
	stock = "<br><span style=\"color:navy\">Last filled <b>STOCK</b> details</span><table style='font-size:80%'><tr><th>Date</th><th>Invoice</th><th>Petrol Bought</th><th>Petrol Cost</th><th>Diesel Bought</th><th>Diesel Cost</th><th>Total Cost</th></tr><tr>" 
	from openpyxl import Workbook, load_workbook
	wb = load_workbook("data/"+dt[:4]+"/"+dt[5:7]+"/"+week+"/data.xlsx")
	ws = wb["addstock"]
	for i in range(1,ws.max_column+1):
		stock = stock + "<td>" + ws.cell(row=ws.max_row,column=i).value + "</td>"
	stock = stock + "</tr></table>"
	
except:
	lastdate = lastdate + "<tr><td>No update</td>"
try:
	f = open("data/salerecovery.txt",'r')
	lines = f.read().splitlines()
	dt = lines[-1]
	week = ""
	val = int(dt[8:10])
	if val <= 7:
		week = "week1"
	elif val >= 8 and val <= 14:
		week = "week2"
	elif val >= 15 and val <= 21:
		week = "week3"
	elif val >= 22:
		week = "week4"
	lastdate = lastdate + "<td>"+dt+"</td>"
	
	sale = "<br><span style=\"color:navy\">Last filled <b>Sale</b> details</span><table style='font-size:80%'><tr><th>Date</th><th>Petrol Sold</th><th>Petrol Test</th><th>Petrol Rate</th><th>Petrol SP</th><th>Diesel Sold</th><th>Diesel Test</th><th>Diesel Rate</th><th>Diesel SP</th><th>Total SP</th></tr><tr>" 
	from openpyxl import Workbook, load_workbook
	wb = load_workbook("data/"+dt[:4]+"/"+dt[5:7]+"/"+week+"/data.xlsx")
	ws = wb["addsale"]
	for i in range(1,ws.max_column+1):
		sale = sale + "<td>" + ws.cell(row=ws.max_row,column=i).value + "</td>"
	sale = sale + "</tr></table>"
except:
	lastdate = lastdate + "<td>No update</td>"
instock = "<br><span style=\"color:navy\">Currently in stock</span><table style='font-size:80%'><tr><th>Petrol</th><th>Diesel</th></tr>"

try:
	f = open("data/cashhandrecovery.txt",'r')
	lines = f.read().splitlines()
	dt = lines[-1]
	week = ""
	val = int(dt[8:10])
	if val <= 7:
		week = "week1"
	elif val >= 8 and val <= 14:
		week = "week2"
	elif val >= 15 and val <= 21:
		week = "week3"
	elif val >= 22:
		week = "week4"
	lastdate = lastdate + "<td>"+dt+"</td></tr></table>"
	
	cash = "<br><span style=\"color:navy\">Last filled <b>CashHand</b> details</span><table style='font-size:80%'><tr><th>Date</th><th>Sold Amount</th><th>PhonePe</th><th>PNB Credit Card</th><th>Fino Payment</th><th>Smart Card</th><th>Account Payment</th><th>Expense Amount</th><th>Udhar Amount</th><th>Total Amount</th><th>Cash Hand</th><th>Deposit</th><th>Locker</th></tr><tr>" 
	from openpyxl import Workbook, load_workbook
	wb = load_workbook("data/"+dt[:4]+"/"+dt[5:7]+"/"+week+"/data.xlsx")
	ws = wb["addcashhand"]
	for i in range(1,ws.max_column+1):
		cash = cash + "<td>" + ws.cell(row=ws.max_row,column=i).value + "</td>"
	cash = cash + "</tr></table>"
except:
	lastdate = lastdate + "<td>No update</td></tr></table>"
instock = "<br><span style=\"color:navy\">Current Status</span><table><tr><th>Petrol</th><th>Diesel</th><th>Cash in locker</th></tr>"

try:
	petrol = open("data/petrol.txt","r")
	petrol = petrol.read()
	petrol = petrol[:-1]
	instock = instock + "<tr><td>" +petrol+ "</td>"
except:
	instock = instock + "<tr><td>0</td>"

try:
	diesel = open("data/diesel.txt","r")
	diesel = diesel.read()
	diesel = diesel[:-1]
	instock = instock + "<td>" +diesel+ "</td>"
except:
	instock = instock + "<td>0</td>"
	
try:
	locker = open("data/locker.txt","r")
	locker = locker.read()
	locker = locker[:-1]
	instock = instock + "<td>" +locker+ "</td></tr></table>"
except:
	instock = instock + "<td>0</td></tr></table>"

print(lastdate+stock+sale+cash+instock)
