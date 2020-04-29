#!/usr/bin/python3
print()

password = "2870"
import cgi
import cgitb
cgitb.enable()
form = cgi.FieldStorage()
passcode = form.getvalue("pass")

if passcode == password:
	dt = form.getvalue("date")
	sa = form.getvalue("sa")
	a1 = form.getvalue("a1")
	a2 = form.getvalue("a2")
	a3 = form.getvalue("a3")
	a4 = form.getvalue("a4")
	a5 = form.getvalue("a5")
	a6 = form.getvalue("a6")
	a7 = form.getvalue("a7")
	ta = form.getvalue("ta")
	ch = form.getvalue("ch")
	d = form.getvalue("d")
	week = ""
	val = int(dt[8:10])
	if val <= 7:
		week = "week1"
	elif val >= 8 and val <= 14:
		week = "week2"
	elif val >= 15 and val <= 21:
		week = "week3"
	elif val >= 22:
		week = "week4"
	
	import os
	
	from openpyxl import Workbook, load_workbook
	if os.path.exists("data/"+dt[:4]) == True:
		if os.path.exists("data/"+dt[:4]+"/"+dt[5:7]) == True:
			if os.path.exists("data/"+dt[:4]+"/"+dt[5:7]+"/"+week) == True:
				locker = open("data/locker.txt","r")
				locker = locker.read()
				locker = locker[:-1]
				locker = float(locker) + float(ch) - float(d)
				locker = round(locker,2)
				locker = str(locker)
				os.system("echo "+locker+" > data/locker.txt")
				
				wb = load_workbook("data/"+dt[:4]+"/"+dt[5:7]+"/"+week+"/data.xlsx")
				wb['addcashhand'].append([dt,sa,a1,a2,a3,a4,a5,a6,a7,ta,ch,d,locker])
				wb.save("data/"+dt[:4]+"/"+dt[5:7]+"/"+week+"/data.xlsx")
				
				wb = load_workbook("data/"+dt[:4]+"/"+dt[5:7]+"/data.xlsx")
				wb['addcashhand'].append([dt,sa,a1,a2,a3,a4,a5,a6,a7,ta,ch,d,locker])
				wb.save("data/"+dt[:4]+"/"+dt[5:7]+"/data.xlsx")
				
				wb = load_workbook("data/"+dt[:4]+"/data.xlsx")
				wb['addcashhand'].append([dt,sa,a1,a2,a3,a4,a5,a6,a7,ta,ch,d,locker])
				wb.save("data/"+dt[:4]+"/data.xlsx")
			else:
				print("No record for this week")
		else:
			print("No record for this month")
	else:
		print("No record for this year")
	
	os.system("echo "+dt+" >> data/cashhandrecovery.txt")
	f = open("data/cashhandrecovery.txt",'r')
	lines = f.read().splitlines()
	if len(lines) > 10:
		st = ""
		for line in lines[1:-1]:
			st = st + line + "\n"
		st = st + lines[-1]
		os.system("echo "+st+" > data/cashhandrecovery.txt")
	
	print("CashHand Data added for date-"+dt)
else:
	print("Wrong PassCode")
