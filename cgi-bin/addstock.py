#!/usr/bin/python3
print()

password = "2870"
import cgi
import cgitb
cgitb.enable()
form = cgi.FieldStorage()
passcode = form.getvalue("pass")

if passcode == password:
	dt = form.getvalue("date")
	iv = form.getvalue("invoice")
	p = form.getvalue("petrol")
	pc = form.getvalue("petrolcost")
	d = form.getvalue("diesel")
	dc = form.getvalue("dieselcost")
	t = form.getvalue("total")
	week = ""
	val = int(dt[8:10])
	if val <= 7:
		week = "week1"
	elif val >= 8 and val <= 14:
		week = "week2"
	elif val >= 15 and val <= 21:
		week = "week3"
	elif val >= 22:
		week = "week4"
	
	import os
	
	from openpyxl import Workbook, load_workbook
	if os.path.exists("data") == False:
		os.mkdir("data")
		os.system("echo 0 > data/petrol.txt")
		os.system("echo 0 > data/diesel.txt")
		os.system("echo 0 > data/locker.txt")
	if os.path.exists("data/"+dt[:4]) == True:
		if os.path.exists("data/"+dt[:4]+"/"+dt[5:7]) == True:
			if os.path.exists("data/"+dt[:4]+"/"+dt[5:7]+"/"+week) == True:
			
				wb = load_workbook("data/"+dt[:4]+"/"+dt[5:7]+"/"+week+"/data.xlsx")
				wb['addstock'].append([dt,iv,p,pc,d,dc,t])
				wb.save("data/"+dt[:4]+"/"+dt[5:7]+"/"+week+"/data.xlsx")
				
				wb = load_workbook("data/"+dt[:4]+"/"+dt[5:7]+"/data.xlsx")
				wb['addstock'].append([dt,iv,p,pc,d,dc,t])
				wb.save("data/"+dt[:4]+"/"+dt[5:7]+"/data.xlsx")
				
				wb = load_workbook("data/"+dt[:4]+"/data.xlsx")
				wb['addstock'].append([dt,iv,p,pc,d,dc,t])
				wb.save("data/"+dt[:4]+"/data.xlsx")
			else:
				os.mkdir("data/"+dt[:4]+"/"+dt[5:7]+"/"+week)
				os.system("cp data/petrol.txt "+"data/"+dt[:4]+"/"+dt[5:7]+"/"+week+"/openpetrol.txt")
				os.system("cp data/diesel.txt "+"data/"+dt[:4]+"/"+dt[5:7]+"/"+week+"/opendiesel.txt")
				wb = Workbook()
				wb.create_sheet("addstock",0)
				wb.create_sheet("addsale",1)
				wb.create_sheet("addcashhand",2)
				wb['addstock'].append(["Date","Invoice","Petrol in liters","Petrol cost","Diesel in liters","Diesel cost","Total cost"])
				wb['addsale'].append(["Date","Petrol in liters","Petrol for testing","Petrol Rate","Petrol selling price","Diesel in liters","Diesel for testing","Diesel Rate","Diesel selling price","Total cost"])
				wb['addcashhand'].append(["Date","Sold Amount","PhonePe","PNB CreditCard","Fino Payment","Smart Card","Account Payment","Expense Amount","Udhar Amount","Total Amount","Cash Hand","Daily Deposit","In Locker"])
				wb['addstock'].append([dt,iv,p,pc,d,dc,t])
				wb.save("data/"+dt[:4]+"/"+dt[5:7]+"/"+week+"/data.xlsx")
		else:
			os.mkdir("data/"+dt[:4]+"/"+dt[5:7])
			os.system("cp data/petrol.txt "+"data/"+dt[:4]+"/"+dt[5:7]+"/openpetrol.txt")
			os.system("cp data/diesel.txt "+"data/"+dt[:4]+"/"+dt[5:7]+"/opendiesel.txt")
			wb = Workbook()
			wb.create_sheet("addstock",0)
			wb.create_sheet("addsale",1)
			wb.create_sheet("addcashhand",2)
			wb['addstock'].append(["Date","Invoice","Petrol in liters","Petrol cost","Diesel in liters","Diesel cost","Total cost"])
			wb['addsale'].append(["Date","Petrol in liters","Petrol for testing","Petrol Rate","Petrol selling price","Diesel in liters","Diesel for testing","Diesel Rate","Diesel selling price","Total cost"])
			wb['addcashhand'].append(["Date","Sold Amount","PhonePe","PNB CreditCard","Fino Payment","Smart Card","Account Payment","Expense Amount","Udhar Amount","Total Amount","Cash Hand","Daily Deposit","In Locker"])
			wb['addstock'].append([dt,iv,p,pc,d,dc,t])
			wb.save("data/"+dt[:4]+"/"+dt[5:7]+"/data.xlsx")
			
			os.mkdir("data/"+dt[:4]+"/"+dt[5:7]+"/"+week)
			os.system("cp data/petrol.txt "+"data/"+dt[:4]+"/"+dt[5:7]+"/"+week+"/openpetrol.txt")
			os.system("cp data/diesel.txt "+"data/"+dt[:4]+"/"+dt[5:7]+"/"+week+"/opendiesel.txt")
			wb = Workbook()
			wb.create_sheet("addstock",0)
			wb.create_sheet("addsale",1)
			wb.create_sheet("addcashhand",2)
			wb['addstock'].append(["Date","Invoice","Petrol in liters","Petrol cost","Diesel in liters","Diesel cost","Total cost"])
			wb['addsale'].append(["Date","Petrol in liters","Petrol for testing","Petrol Rate","Petrol selling price","Diesel in liters","Diesel for testing","Diesel Rate","Diesel selling price","Total cost"])
			wb['addcashhand'].append(["Date","Sold Amount","PhonePe","PNB CreditCard","Fino Payment","Smart Card","Account Payment","Expense Amount","Udhar Amount","Total Amount","Cash Hand","Daily Deposit","In Locker"])
			wb['addstock'].append([dt,iv,p,pc,d,dc,t])
			wb.save("data/"+dt[:4]+"/"+dt[5:7]+"/"+week+"/data.xlsx")
	else:
		os.mkdir("data/"+dt[:4])
		os.system("cp data/petrol.txt "+"data/"+dt[:4]+"/openpetrol.txt")
		os.system("cp data/diesel.txt "+"data/"+dt[:4]+"/opendiesel.txt")
		wb = Workbook()
		wb.create_sheet("addstock",0)
		wb.create_sheet("addsale",1)
		wb.create_sheet("addcashhand",2)
		wb['addstock'].append(["Date","Invoice","Petrol in liters","Petrol cost","Diesel in liters","Diesel cost","Total cost"])
		wb['addsale'].append(["Date","Petrol in liters","Petrol for testing","Petrol Rate","Petrol selling price","Diesel in liters","Diesel for testing","Diesel Rate","Diesel selling price","Total cost"])
		wb['addcashhand'].append(["Date","Sold Amount","PhonePe","PNB CreditCard","Fino Payment","Smart Card","Account Payment","Expense Amount","Udhar Amount","Total Amount","Cash Hand","Daily Deposit","In Locker"])
		wb['addstock'].append([dt,iv,p,pc,d,dc,t])
		wb.save("data/"+dt[:4]+"/data.xlsx")
		
		os.mkdir("data/"+dt[:4]+"/"+dt[5:7])
		os.system("cp data/petrol.txt "+"data/"+dt[:4]+"/"+dt[5:7]+"/openpetrol.txt")
		os.system("cp data/diesel.txt "+"data/"+dt[:4]+"/"+dt[5:7]+"/opendiesel.txt")
		wb = Workbook()
		wb.create_sheet("addstock",0)
		wb.create_sheet("addsale",1)
		wb.create_sheet("addcashhand",2)
		wb['addstock'].append(["Date","Invoice","Petrol in liters","Petrol cost","Diesel in liters","Diesel cost","Total cost"])
		wb['addsale'].append(["Date","Petrol in liters","Petrol for testing","Petrol Rate","Petrol selling price","Diesel in liters","Diesel for testing","Diesel Rate","Diesel selling price","Total cost"])
		wb['addcashhand'].append(["Date","Sold Amount","PhonePe","PNB CreditCard","Fino Payment","Smart Card","Account Payment","Expense Amount","Udhar Amount","Total Amount","Cash Hand","Daily Deposit","In Locker"])
		wb['addstock'].append([dt,iv,p,pc,d,dc,t])
		wb.save("data/"+dt[:4]+"/"+dt[5:7]+"/data.xlsx")
		
		os.mkdir("data/"+dt[:4]+"/"+dt[5:7]+"/"+week)
		os.system("cp data/petrol.txt "+"data/"+dt[:4]+"/"+dt[5:7]+"/"+week+"/openpetrol.txt")
		os.system("cp data/diesel.txt "+"data/"+dt[:4]+"/"+dt[5:7]+"/"+week+"/opendiesel.txt")
		wb = Workbook()
		wb.create_sheet("addstock",0)
		wb.create_sheet("addsale",1)
		wb.create_sheet("addcashhand",2)
		wb['addstock'].append(["Date","Invoice","Petrol in liters","Petrol cost","Diesel in liters","Diesel cost","Total cost"])
		wb['addsale'].append(["Date","Petrol in liters","Petrol for testing","Petrol Rate","Petrol selling price","Diesel in liters","Diesel for testing","Diesel Rate","Diesel selling price","Total cost"])
		wb['addcashhand'].append(["Date","Sold Amount","PhonePe","PNB CreditCard","Fino Payment","Smart Card","Account Payment","Expense Amount","Udhar Amount","Total Amount","Cash Hand","Daily Deposit","In Locker"])
		wb['addstock'].append([dt,iv,p,pc,d,dc,t])
		wb.save("data/"+dt[:4]+"/"+dt[5:7]+"/"+week+"/data.xlsx")
	
	os.system("echo "+dt+" >> data/stockrecovery.txt")
	f = open("data/stockrecovery.txt",'r')
	lines = f.read().splitlines()
	if len(lines) > 10:
		st = ""
		for line in lines[1:-1]:
			st = st + line + "\n"
		st = st + lines[-1]
		os.system("echo "+st+" > data/stockrecovery.txt")
	
	if os.path.exists("data/petrol.txt"):
		petrol = open("data/petrol.txt","r")
		petrol = petrol.read()
		petrol = petrol[:-1]
		petrol = float(petrol) + float(p)
		petrol = str(petrol)
		os.system("echo "+petrol+" > data/petrol.txt")
	else:
		os.system("echo "+p+" > data/petrol.txt")
	
	if os.path.exists("data/diesel.txt"):
		diesel = open("data/diesel.txt","r")
		diesel = diesel.read()
		diesel = diesel[:-1]
		diesel = float(diesel) + float(d)
		diesel = str(diesel)
		os.system("echo "+diesel+" > data/diesel.txt")
	else:
		os.system("echo "+d+" > data/diesel.txt")
	
	print("Stock added for date-"+dt)
else:
	print("Wrong PassCode")
