from openpyxl import Workbook, load_workbook
wb = Workbook()
#wb.create_sheet("addstock",0)
wb.create_sheet("report",1)
try:
	wb.remove(wb["addstock"])
except:
	print()

#wb["addstock"]
