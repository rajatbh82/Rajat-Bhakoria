#!/usr/bin/python3
print()


import cgi
import cgitb
cgitb.enable()
form = cgi.FieldStorage()
dt = form.getvalue("date")
try:
	week = ""
	val = int(dt[8:10])
	if val <= 7:
		week = "week1"
	elif val >= 8 and val <= 14:
		week = "week2"
	elif val >= 15 and val <= 21:
		week = "week3"
	elif val >= 22:
		week = "week4"
	from openpyxl import load_workbook
	wb = load_workbook("data/"+dt[:4]+"/"+dt[5:7]+"/"+week+"/data.xlsx")
	ws = wb["addsale"]
	flag = 0
	for i in range(2,ws.max_row+1):
		if ws.cell(row=i,column=1).value == dt:
			print(ws.cell(row=i,column=ws.max_column).value)
			flag = 1
	if flag == 0:
		print("First fill sale for the given date")
except:
	print("No data found for the given date")
